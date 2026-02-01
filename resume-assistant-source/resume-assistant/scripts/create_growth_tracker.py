#!/usr/bin/env python3
"""
Generate growth tracking spreadsheet from ability improvement plan.

Usage:
    python create_growth_tracker.py --plan growth_plan.json --output tracker.xlsx

Generates an Excel spreadsheet with:
- Weekly task checklist
- Progress tracking
- Milestone checkpoints
- Resource links
"""

import json
import argparse
import sys
from pathlib import Path
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def create_growth_tracker(plan_data: dict, output_path: str) -> None:
    """
    Create an Excel tracking spreadsheet from growth plan data.

    Args:
        plan_data: Growth plan dictionary
        output_path: Output Excel file path
    """
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets
    create_overview_sheet(wb, plan_data)
    create_weekly_tracker_sheet(wb, plan_data)
    create_milestones_sheet(wb, plan_data)
    create_resources_sheet(wb, plan_data)

    # Save workbook
    wb.save(output_path)
    print(f"✅ Growth tracker created: {output_path}")
    print(f"📊 Includes: Overview, Weekly Tasks, Milestones, Resources")
    print(f"💡 Open in Excel/WPS/Numbers to start tracking!")


def create_overview_sheet(wb: openpyxl.Workbook, plan_data: dict):
    """Create overview sheet with plan summary."""
    ws = wb.create_sheet("总览", 0)

    # Styles
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(color="FFFFFF", bold=True)

    # Title
    ws['A1'] = '能力提升计划 - 总览'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')

    # Basic info
    row = 3
    ws[f'A{row}'] = '目标职位'
    ws[f'B{row}'] = plan_data.get('target_position', '')
    ws[f'A{row}'].font = header_font
    row += 1

    ws[f'A{row}'] = '计划时长'
    ws[f'B{row}'] = f"{plan_data.get('duration_weeks', 0)} 周"
    ws[f'A{row}'].font = header_font
    row += 1

    ws[f'A{row}'] = '开始日期'
    ws[f'B{row}'] = plan_data.get('start_date', datetime.now().strftime('%Y-%m-%d'))
    ws[f'A{row}'].font = header_font
    row += 2

    # Current match rate
    ws[f'A{row}'] = '当前匹配度'
    ws[f'B{row}'] = f"{plan_data.get('current_match', 0)}/10"
    ws[f'A{row}'].font = header_font
    row += 1

    ws[f'A{row}'] = '目标匹配度'
    ws[f'B{row}'] = f"{plan_data.get('target_match', 8)}/10"
    ws[f'A{row}'].font = header_font
    row += 2

    # Phases
    ws[f'A{row}'] = '阶段规划'
    ws[f'A{row}'].font = header_font
    row += 1

    phases = plan_data.get('phases', [])
    for i, phase in enumerate(phases, 1):
        ws[f'A{row}'] = f"阶段{i}"
        ws[f'B{row}'] = phase.get('name', '')
        ws[f'C{row}'] = f"第{phase.get('weeks', '')}周"
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15


def create_weekly_tracker_sheet(wb: openpyxl.Workbook, plan_data: dict):
    """Create weekly task tracking sheet."""
    ws = wb.create_sheet("每周任务")

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['周次', '阶段', '任务', '预计时间', '完成状态', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Tasks
    row = 2
    phases = plan_data.get('phases', [])

    for phase in phases:
        phase_name = phase.get('name', '')
        weeks_range = phase.get('weeks', '1-4')

        for task in phase.get('tasks', []):
            ws.cell(row=row, column=1, value=weeks_range)
            ws.cell(row=row, column=2, value=phase_name)
            ws.cell(row=row, column=3, value=task.get('name', ''))
            ws.cell(row=row, column=4, value=f"{task.get('hours', 0)}小时/周")
            ws.cell(row=row, column=5, value='☐ 未完成')
            ws.cell(row=row, column=6, value='')

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border

            row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30

    # Add instructions
    row += 2
    ws.cell(row=row, column=1, value='使用说明：')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='1. 完成任务后，在"完成状态"列改为 ✓ 已完成')
    row += 1
    ws.cell(row=row, column=1, value='2. 在"备注"列记录遇到的问题或心得')
    row += 1
    ws.cell(row=row, column=1, value='3. 建议每周日回顾并更新进度')


def create_milestones_sheet(wb: openpyxl.Workbook, plan_data: dict):
    """Create milestone checkpoints sheet."""
    ws = wb.create_sheet("里程碑")

    # Styles
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['时间点', '里程碑', '检验标准', '达成状态', '实际日期']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Milestones
    row = 2
    milestones = plan_data.get('milestones', [])

    for milestone in milestones:
        ws.cell(row=row, column=1, value=f"第{milestone.get('week', '')}周")
        ws.cell(row=row, column=2, value=milestone.get('goal', ''))
        ws.cell(row=row, column=3, value=milestone.get('criteria', ''))
        ws.cell(row=row, column=4, value='☐ 未达成')
        ws.cell(row=row, column=5, value='')

        # Apply borders
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15


def create_resources_sheet(wb: openpyxl.Workbook, plan_data: dict):
    """Create learning resources sheet."""
    ws = wb.create_sheet("学习资源")

    # Styles
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['技能/领域', '资源类型', '资源名称', '链接/说明', '优先级']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Resources
    row = 2
    resources = plan_data.get('resources', [])

    for resource in resources:
        ws.cell(row=row, column=1, value=resource.get('skill', ''))
        ws.cell(row=row, column=2, value=resource.get('type', ''))
        ws.cell(row=row, column=3, value=resource.get('name', ''))
        ws.cell(row=row, column=4, value=resource.get('link', ''))
        ws.cell(row=row, column=5, value=resource.get('priority', ''))

        # Apply borders
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 10


def main():
    parser = argparse.ArgumentParser(description="Generate growth tracking spreadsheet")
    parser.add_argument("--plan", "-p", required=True, help="JSON file with growth plan data")
    parser.add_argument("--output", "-o", default="growth_tracker.xlsx", help="Output Excel file path")

    args = parser.parse_args()

    # Load plan data
    plan_path = Path(args.plan)
    if not plan_path.exists():
        print(f"Error: Plan file not found: {args.plan}")
        sys.exit(1)

    with open(plan_path, 'r', encoding='utf-8') as f:
        try:
            plan_data = json.load(f)
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON in {args.plan}: {e}")
            sys.exit(1)

    create_growth_tracker(plan_data, args.output)


if __name__ == "__main__":
    main()
