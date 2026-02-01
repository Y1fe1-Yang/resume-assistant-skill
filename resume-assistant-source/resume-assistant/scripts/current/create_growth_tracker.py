#!/usr/bin/env python3
"""
Generate growth tracking spreadsheet from ability improvement plan.

Usage:
    python create_growth_tracker.py --plan growth_plan.json --output tracker.xlsx

Generates an Excel spreadsheet with:
- Weekly task checklist
- Progress tracking
- Milestone checkpoints
- Resource links
"""

import json
import argparse
import sys
from pathlib import Path
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def create_growth_tracker(plan_data: dict, output_path: str) -> None:
    """
    Create an Excel tracking spreadsheet from growth plan data.

    Args:
        plan_data: Growth plan dictionary
        output_path: Output Excel file path
    """
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets
    create_overview_sheet(wb, plan_data)
    create_weekly_tracker_sheet(wb, plan_data)
    create_milestones_sheet(wb, plan_data)
    create_resources_sheet(wb, plan_data)

    # Save workbook
    wb.save(output_path)
    print(f"✅ Growth tracker created: {output_path}")
    print(f"📊 Includes: Overview, Weekly Tasks, Milestones, Resources")
    print(f"💡 Open in Excel/WPS/Numbers to start tracking!")


def create_overview_sheet(wb: openpyxl.Workbook, plan_data: dict):
    """Create overview sheet with plan summary."""
    ws = wb.create_sheet("总览", 0)

    # Styles
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(color="FFFFFF", bold=True)

    # Title
    ws['A1'] = '能力提升计划 - 总览'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')

    # Parse timeline (支持"3个月"、"12周"等格式)
    timeline = plan_data.get('timeline', '')
    duration_weeks = 12  # default
    if '个月' in timeline:
        months = int(''.join(filter(str.isdigit, timeline)))
        duration_weeks = months * 4
    elif '周' in timeline:
        duration_weeks = int(''.join(filter(str.isdigit, timeline)))

    # Basic info
    row = 3
    ws[f'A{row}'] = '目标职位'
    ws[f'B{row}'] = plan_data.get('target_position', '')
    ws[f'A{row}'].font = header_font
    row += 1

    ws[f'A{row}'] = '计划时长'
    ws[f'B{row}'] = f"{duration_weeks} 周"
    ws[f'A{row}'].font = header_font
    row += 1

    ws[f'A{row}'] = '开始日期'
    ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d')
    ws[f'A{row}'].font = header_font
    row += 2

    # Current match rate
    ws[f'A{row}'] = '当前匹配度'
    ws[f'B{row}'] = f"0/10"
    ws[f'A{row}'].font = header_font
    row += 1

    ws[f'A{row}'] = '目标匹配度'
    ws[f'B{row}'] = f"8/10"
    ws[f'A{row}'].font = header_font
    row += 2

    # Phases
    ws[f'A{row}'] = '阶段规划'
    ws[f'A{row}'].font = header_font
    row += 1

    phases = plan_data.get('phases', [])
    for i, phase in enumerate(phases, 1):
        ws[f'A{row}'] = f"阶段{i}"
        # 支持新格式的phase/title字段
        phase_name = phase.get('title', phase.get('phase', ''))
        ws[f'B{row}'] = phase_name
        # 根据阶段数量计算周数
        weeks_per_phase = duration_weeks // len(phases) if phases else 4
        start_week = (i-1) * weeks_per_phase + 1
        end_week = i * weeks_per_phase
        ws[f'C{row}'] = f"第{start_week}-{end_week}周"
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15


def create_weekly_tracker_sheet(wb: openpyxl.Workbook, plan_data: dict):
    """Create weekly task tracking sheet with detailed daily breakdowns."""
    ws = wb.create_sheet("每周任务")

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    phase_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    task_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['周次', '阶段', '主任务', '具体行动项', '预计工时', '截止日期', '完成状态', '实际用时', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Tasks
    row = 2
    phases = plan_data.get('phases', [])

    # Parse timeline for week calculation
    timeline = plan_data.get('timeline', '12周')
    duration_weeks = 12
    if '个月' in timeline:
        months = int(''.join(filter(str.isdigit, timeline)))
        duration_weeks = months * 4
    elif '周' in timeline:
        duration_weeks = int(''.join(filter(str.isdigit, timeline)))

    weeks_per_phase = duration_weeks // len(phases) if phases else 4

    for phase_idx, phase in enumerate(phases, 1):
        # 支持新格式
        phase_name = phase.get('title', phase.get('phase', ''))
        start_week = (phase_idx-1) * weeks_per_phase + 1
        end_week = phase_idx * weeks_per_phase

        # 支持新格式的tasks
        for task in phase.get('tasks', []):
            task_name = task.get('task', task.get('name', ''))
            deadline = task.get('deadline', '')

            # 分解任务为具体行动项
            subtasks = _break_down_task(task_name, weeks_per_phase)

            for week_offset, subtask_info in enumerate(subtasks):
                current_week = start_week + week_offset
                if current_week > end_week:
                    break

                ws.cell(row=row, column=1, value=f"第{current_week}周")
                ws.cell(row=row, column=2, value=phase_name if week_offset == 0 else '')
                ws.cell(row=row, column=3, value=task_name if week_offset == 0 else '')
                ws.cell(row=row, column=4, value=subtask_info['action'])
                ws.cell(row=row, column=5, value=subtask_info['hours'])
                ws.cell(row=row, column=6, value=subtask_info['deadline'])
                ws.cell(row=row, column=7, value='☐ 未完成')
                ws.cell(row=row, column=8, value='')
                ws.cell(row=row, column=9, value='')

                # Apply styling
                if week_offset == 0:
                    ws.cell(row=row, column=2).fill = phase_fill
                    ws.cell(row=row, column=3).fill = task_fill

                # Apply borders
                for col in range(1, 10):
                    ws.cell(row=row, column=col).border = border

                row += 1

            # 添加空行分隔不同任务
            row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 30

    # Add instructions
    row += 2
    ws.cell(row=row, column=1, value='使用说明：')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='1. 每天完成具体行动项后，在"完成状态"列改为 ✓ 已完成，并记录实际用时')
    row += 1
    ws.cell(row=row, column=1, value='2. 在"备注"列记录学习要点、遇到的问题、解决方案')
    row += 1
    ws.cell(row=row, column=1, value='3. 预计工时仅供参考，根据实际情况调整')
    row += 1
    ws.cell(row=row, column=1, value='4. 建议每周日回顾本周进度，规划下周任务')


def _break_down_task(task_name: str, weeks_available: int) -> list:
    """
    将主任务分解为具体的周度行动项。

    Args:
        task_name: 主任务名称
        weeks_available: 可用的周数

    Returns:
        包含具体行动项的列表，每项包含action、hours、deadline
    """
    subtasks = []

    # 根据任务类型进行智能分解
    task_lower = task_name.lower()

    # 学习类任务
    if any(kw in task_lower for kw in ['学习', '掌握', '了解', '课程', 'learn']):
        if weeks_available >= 3:
            subtasks = [
                {'action': '📚 理论学习：观看课程前1/3内容，做笔记（每天1-2小时）',
                 'hours': '7-10h', 'deadline': '本周五'},
                {'action': '💻 实践练习：完成配套练习题，搭建基础环境',
                 'hours': '8-12h', 'deadline': '下周三'},
                {'action': '🔧 项目实战：参照教程完成1个小项目，理解核心概念',
                 'hours': '10-15h', 'deadline': '第3周日'},
            ]
        elif weeks_available == 2:
            subtasks = [
                {'action': '📚 集中学习：完整观看课程视频，整理核心知识点',
                 'hours': '10-15h', 'deadline': '本周日'},
                {'action': '💻 实战练习：完成至少3个练习案例，建立肌肉记忆',
                 'hours': '8-12h', 'deadline': '下周日'},
            ]
        else:
            subtasks = [
                {'action': '⚡ 快速上手：观看核心章节，完成基础练习',
                 'hours': '10-15h', 'deadline': '本周日'},
            ]

    # 项目类任务
    elif any(kw in task_lower for kw in ['项目', '开发', '实现', '构建', 'project', 'build']):
        if weeks_available >= 3:
            subtasks = [
                {'action': '📋 需求分析与设计：确定功能范围，画出架构图和流程图',
                 'hours': '4-6h', 'deadline': '本周三'},
                {'action': '🏗️ 核心功能开发：实现主要业务逻辑（MVP版本）',
                 'hours': '12-16h', 'deadline': '第2周日'},
                {'action': '✨ 完善与优化：添加边界处理、错误提示、UI优化',
                 'hours': '8-10h', 'deadline': '第3周五'},
                {'action': '📝 文档与部署：编写README、测试文档，部署上线',
                 'hours': '4-6h', 'deadline': '第3周日'},
            ]
        elif weeks_available == 2:
            subtasks = [
                {'action': '📋 设计与搭建：确定技术栈，搭建项目框架',
                 'hours': '6-8h', 'deadline': '本周五'},
                {'action': '🏗️ 功能实现：完成核心功能开发和基础测试',
                 'hours': '12-15h', 'deadline': '下周日'},
            ]
        else:
            subtasks = [
                {'action': '⚡ 快速搭建：参考现有项目，实现核心demo',
                 'hours': '10-12h', 'deadline': '本周日'},
            ]

    # 阅读类任务
    elif any(kw in task_lower for kw in ['阅读', '研读', '文档', '书籍', 'read', 'book']):
        if weeks_available >= 2:
            subtasks = [
                {'action': '📖 通读全书：每天30-60分钟，完成第一遍阅读',
                 'hours': '8-10h', 'deadline': '第1周日'},
                {'action': '✍️ 精读与笔记：重点章节做详细笔记，整理思维导图',
                 'hours': '6-8h', 'deadline': '第2周日'},
            ]
        else:
            subtasks = [
                {'action': '📖 重点阅读：聚焦核心章节，提炼关键知识点',
                 'hours': '8-10h', 'deadline': '本周日'},
            ]

    # 练习类任务
    elif any(kw in task_lower for kw in ['练习', '刷题', '题目', 'practice', 'exercise']):
        subtasks = [
            {'action': '🎯 基础题（Easy）：每天2-3题，熟悉基本概念',
             'hours': '5-7h', 'deadline': '本周日'},
            {'action': '🎯 进阶题（Medium）：每天1-2题，提升解题能力',
             'hours': '6-8h', 'deadline': '下周日'},
        ]

    # 默认通用分解
    else:
        if weeks_available >= 2:
            subtasks = [
                {'action': f'🚀 启动阶段：{task_name} - 准备工作和基础搭建',
                 'hours': '6-8h', 'deadline': '本周日'},
                {'action': f'⚡ 执行阶段：{task_name} - 核心工作完成',
                 'hours': '8-12h', 'deadline': '下周日'},
            ]
        else:
            subtasks = [
                {'action': f'⚡ {task_name} - 集中完成',
                 'hours': '10-15h', 'deadline': '本周日'},
            ]

    return subtasks


def create_milestones_sheet(wb: openpyxl.Workbook, plan_data: dict):
    """Create milestone checkpoints sheet."""
    ws = wb.create_sheet("里程碑")

    # Styles
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['时间点', '里程碑', '检验标准', '达成状态', '实际日期']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Milestones - 从phases中提取milestone
    row = 2
    phases = plan_data.get('phases', [])

    # Parse timeline
    timeline = plan_data.get('timeline', '12周')
    duration_weeks = 12
    if '个月' in timeline:
        months = int(''.join(filter(str.isdigit, timeline)))
        duration_weeks = months * 4
    elif '周' in timeline:
        duration_weeks = int(''.join(filter(str.isdigit, timeline)))

    weeks_per_phase = duration_weeks // len(phases) if phases else 4

    for phase_idx, phase in enumerate(phases, 1):
        milestone_text = phase.get('milestone', '')
        if milestone_text:
            week_num = phase_idx * weeks_per_phase
            phase_name = phase.get('title', phase.get('phase', ''))

            ws.cell(row=row, column=1, value=f"第{week_num}周")
            ws.cell(row=row, column=2, value=phase_name)
            ws.cell(row=row, column=3, value=milestone_text)
            ws.cell(row=row, column=4, value='☐ 未达成')
            ws.cell(row=row, column=5, value='')

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border

            row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15


def create_resources_sheet(wb: openpyxl.Workbook, plan_data: dict):
    """Create learning resources sheet."""
    ws = wb.create_sheet("学习资源")

    # Styles
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['阶段', '任务', '资源名称', '类型', '优先级']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Resources - 从phases的tasks中提取resources
    row = 2
    phases = plan_data.get('phases', [])

    for phase in phases:
        phase_name = phase.get('title', phase.get('phase', ''))
        tasks = phase.get('tasks', [])

        for task in tasks:
            task_name = task.get('task', task.get('name', ''))
            resources = task.get('resources', [])

            for resource in resources:
                ws.cell(row=row, column=1, value=phase_name)
                ws.cell(row=row, column=2, value=task_name)
                ws.cell(row=row, column=3, value=resource)
                # 根据资源名称推断类型
                resource_type = '在线课程' if 'Udemy' in resource or 'Coursera' in resource else \
                               '书籍' if '《' in resource else \
                               '视频' if 'B站' in resource or 'YouTube' in resource else \
                               '文档'
                ws.cell(row=row, column=4, value=resource_type)
                ws.cell(row=row, column=5, value='高')

                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border

                row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10


def main():
    parser = argparse.ArgumentParser(description="Generate growth tracking spreadsheet")
    parser.add_argument("--plan", "-p", required=True, help="JSON file with growth plan data")
    parser.add_argument("--output", "-o", default="growth_tracker.xlsx", help="Output Excel file path")

    args = parser.parse_args()

    # Load plan data
    plan_path = Path(args.plan)
    if not plan_path.exists():
        print(f"Error: Plan file not found: {args.plan}")
        sys.exit(1)

    with open(plan_path, 'r', encoding='utf-8') as f:
        try:
            plan_data = json.load(f)
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON in {args.plan}: {e}")
            sys.exit(1)

    create_growth_tracker(plan_data, args.output)


if __name__ == "__main__":
    main()
